#!/usr/bin/env python3
print("Content-Type: text/html\n")

import os
import shutil
import xml.etree.ElementTree as ET
import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook
from db_connection import get_connection
import csv

import ftplib
import logging
import time
from ftp_config import FTP_CONFIGS

# =======================================================
# Time Calculation
# =======================================================
def format_duration(seconds):
    mins = int(seconds // 60)
    secs = int(seconds % 60)
    ms = int((seconds - int(seconds)) * 1000)
    return f"{mins} min {secs} sec {ms} ms"

# =======================================================
# BATCH CONFIG
# =======================================================

BATCH_SIZE = 5000

# ===========================================
# SELECT LOG DIRECTORY BASED ON FTP TYPE
# ===========================================
def get_log_dir(type):
    CFG = FTP_CONFIGS["GLOBAL"]

    if type == "WESTSIDE_WS":
        return CFG["ws_local_log"]
    elif type == "WESTSIDE_TUL":
        return CFG["tul_local_log"]
    
# =======================================================
# LOGGING SETUP
# =======================================================
    
def setup_logging(type):
    LOG_DIR = get_log_dir(type)
    os.makedirs(LOG_DIR, exist_ok=True)

    today = time.strftime("%Y-%m-%d")
    LOG_FILE = os.path.join(LOG_DIR, f"{type.lower()}_sales_order_{today}.log")

    logging.basicConfig(
        filename=LOG_FILE,
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s"
    )

    logging.info(f"--------------- {type} SALES ORDER IMPORT STARTED ---------------")



# =======================================================
# FTP DOWNLOAD (XML/XLSX ONLY)
# =======================================================
def ftp_download_files(type):

    setup_logging(type)

    logging.info("[FTP] Connecting to FTP Server...")
    
    FTPDIR = FTP_CONFIGS["GLOBAL"]

    if type == "WESTSIDE_WS":
        CFG = FTP_CONFIGS["WESTSIDE_WS"]
        INPUT_FOLDER = FTPDIR["ws_local_read"]
    elif type == "WESTSIDE_TUL":
        CFG = FTP_CONFIGS["WESTSIDE_TUL"]
        INPUT_FOLDER = FTPDIR["tul_local_read"]
    else:
        logging.error(f"[FTP] Unknown type: {type}")
        return []
    

    FTP_HOST = CFG["host"]
    FTP_USER = CFG["user"]
    FTP_PASS = CFG["pass"]

    REMOTE_INCOMING = CFG["remote_incoming"]
    REMOTE_ARCHIVE = CFG["remote_archive"]
    
    try:
        ftp = ftplib.FTP()
        ftp.connect(FTP_HOST, 21, timeout=60)
        ftp.login(FTP_USER, FTP_PASS)
    except Exception as e:
        logging.error(f"[FTP] Connect/login failed: {e}")
        return []

    try:
        ftp.cwd(REMOTE_INCOMING)
        files = ftp.nlst()
    except Exception as e:
        logging.error(f"[FTP] Unable to list files: {e}")
        ftp.quit()
        return []

    downloaded = []
    for f in files:
        if not (f.lower().endswith(".xml") or f.lower().endswith(".xlsx")):
            continue

        local_path = os.path.join(INPUT_FOLDER, f)
        os.makedirs(INPUT_FOLDER, exist_ok=True)

        try:
            with open(local_path, "wb") as fp:
                ftp.retrbinary("RETR " + f, fp.write)
            ftp.rename(f"{REMOTE_INCOMING}/{f}", f"{REMOTE_ARCHIVE}/{f}")
            downloaded.append(f)
        except Exception as e:
            logging.error(f"[FTP ERROR] {f}: {e}")

    ftp.quit()
    return downloaded


# =======================================================
# BATCH INSERT WRAPPER
# =======================================================
def insert_batch_and_count(cursor, insert_query, batch_data):
    if not batch_data:
        return 0
    n = len(batch_data)
    execute_values(cursor, insert_query, batch_data)
    batch_data.clear()
    return n


# =======================================================
# PROCESS XML
# =======================================================
def process_xml(file_path, cursor, type):

    setup_logging(type)

    logging.info(f"[XML] Processing: {file_path}")
    tree = ET.parse(file_path)
    root = tree.getroot()

    batch_data = []
    total = 0

    query = """
        INSERT INTO sales_orders (sku, sale_qty, store_code, consolidate, is_read)
        VALUES %s
    """

    for order in root.findall("OrderLine"):
        sku_el = order.find("ArticleNumber")
        store_el = order.find("P1SlaveId")

        if sku_el is None:
            continue

        sku = sku_el.text.strip()
        store_code = store_el.text.strip() if store_el is not None else ""
        qty = 1

        if not sku:
            continue

        batch_data.append((sku, qty, store_code, 0, 0))

        if len(batch_data) >= BATCH_SIZE:
            total += insert_batch_and_count(cursor, query, batch_data)

    if batch_data:
        total += insert_batch_and_count(cursor, query, batch_data)

    logging.info(f"[XML] Total inserted: {total}")
    return total


# =======================================================
# PROCESS XLSX
# =======================================================
def process_xlsx(file_path, cursor, type):

    setup_logging(type)

    logging.info(f"[XLSX] Processing: {file_path}")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.active

    headers = {name: idx for idx, name in enumerate(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))}

    if "Material" not in headers or "Slave ID" not in headers:
        raise Exception("Required headers missing: Material, Slave ID")

    mat_col = headers["Material"]
    slave_col = headers["Slave ID"]

    batch_data = []
    total = 0

    query = """
        INSERT INTO sales_orders (sku, sale_qty, store_code, consolidate, is_read)
        VALUES %s
    """

    for row in sheet.iter_rows(min_row=2, values_only=True):
        sku = row[mat_col]
        store_code = row[slave_col]

        if not sku:
            continue

        batch_data.append((str(sku), 1, str(store_code), 0, 0))

        if len(batch_data) >= BATCH_SIZE:
            total += insert_batch_and_count(cursor, query, batch_data)

    if batch_data:
        total += insert_batch_and_count(cursor, query, batch_data)

    wb.close()
    logging.info(f"[XLSX] Total inserted: {total}")
    return total

# =======================================================
# MAIN PROCESS
# =======================================================
def orderProcess(type):

    setup_logging(type)
    start_time = time.time()

    FTPDIR = FTP_CONFIGS["GLOBAL"]

    # Always set folders (NO TRY except here)
    if type == "WESTSIDE_WS":
        INPUT_FOLDER = FTPDIR["ws_local_read"]
        ARCHIVE_FOLDER = FTPDIR["ws_local_archive"]

    elif type == "WESTSIDE_TUL":
        INPUT_FOLDER = FTPDIR["tul_local_read"]
        ARCHIVE_FOLDER = FTPDIR["tul_local_archive"]

    else:
        logging.error(f"[ERROR] Unknown type received: {type}")
        return

    # Ensure folders always exist
    os.makedirs(INPUT_FOLDER, exist_ok=True)
    os.makedirs(ARCHIVE_FOLDER, exist_ok=True)

    # FTP download inside safe try
    try:
        ftp_download_files(type)
    except Exception as e:
        logging.error(f"[FTP DOWNLOAD ERROR] {e}")

    conn = get_connection()
    if not conn:
        print("DB connection failed<br>")
        return

    cursor = conn.cursor()

    files = [
        f for f in os.listdir(INPUT_FOLDER)
        if f.lower().endswith((".xml", ".xlsx"))
    ]

    if not files:
        print("No sales order files found<br>")
        return

    for f in files:
        path = os.path.join(INPUT_FOLDER, f)
        print(f"Processing File: {path} <br>")

        try:
            if f.lower().endswith(".xml"):
                count = process_xml(path, cursor, type)
            elif f.lower().endswith(".xlsx"):
                count = process_xlsx(path, cursor, type)

            conn.commit()
            print(f"Inserted rows: {count} <br>")

            shutil.move(path, os.path.join(ARCHIVE_FOLDER, f))
            print(f"Archived: {f} <br>")

        except Exception as e:
            conn.rollback()
            print(f"Error in {f}: {e} <br>")

    cursor.close()
    conn.close()
    print("All files processed<br>")

    # ADD END TIMER + LOG
    end_time = time.time()
    duration = end_time - start_time
    logging.info(f"[TIME] Total Duration for {type}: {format_duration(duration)}")

# =======================================================
# Process Sales Orders Store Procedure
# =======================================================
def salesOrderProcedure():
    conn = get_connection()
    if not conn:
        print("DB connection failed<br>")
        return

    cursor = conn.cursor()

    try:
        cursor.execute("CALL process_sales_orders()")
        conn.commit()
        print("Sales Order Procedure executed successfully<br>")
    except Exception as e:
        conn.rollback()
        print(f"Procedure error: {e}")

    cursor.close()
    conn.close()

orderProcess("WESTSIDE_WS")
orderProcess("WESTSIDE_TUL")
salesOrderProcedure()

