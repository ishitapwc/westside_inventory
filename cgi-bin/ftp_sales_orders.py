#!/usr/bin/env python3
print("Content-Type: text/html\n")

import os
import shutil
import xml.etree.ElementTree as ET
import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook
from db_connection import get_connection

import ftplib
import logging
import time
from ftp_config import FTP_CONFIGS


# =======================================================
# FTP CONFIG (WESTSIDE_2)
# =======================================================
CFG = FTP_CONFIGS["WESTSIDE_2"]

FTP_HOST = CFG["host"]
FTP_USER = CFG["user"]
FTP_PASS = CFG["pass"]

REMOTE_INCOMING = CFG["remote_incoming"]
REMOTE_ARCHIVE = CFG["remote_archive"]

INPUT_FOLDER = CFG["local_read"]          # Local read folder
ARCHIVE_FOLDER = CFG["local_archive"]     # Local archive folder
LOG_DIR = CFG["local_log"]                # Local logs folder

BATCH_SIZE = 5000


# =======================================================
# LOGGING SETUP
# =======================================================
os.makedirs(LOG_DIR, exist_ok=True)

today = time.strftime("%Y-%m-%d")
LOG_FILE = os.path.join(LOG_DIR, f"sales_order_import_{today}.log")

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)

logging.info("--------------- SALES ORDER IMPORT STARTED ---------------")


# =======================================================
# FTP DOWNLOAD FUNCTION (downloads only xml/xlsx)
# =======================================================
def ftp_download_files():
    logging.info("[FTP] Connecting to FTP Server...")

    try:
        ftp = ftplib.FTP()
        ftp.connect(FTP_HOST, 21, timeout=60)
        ftp.login(FTP_USER, FTP_PASS)
    except Exception as e:
        logging.error(f"[FTP] Connect/login failed: {e}")
        return []

    try:
        ftp.cwd(REMOTE_INCOMING)
    except Exception as e:
        logging.warning(f"[FTP] Could not cwd to {REMOTE_INCOMING}: {e}")

    try:
        files = ftp.nlst()
    except Exception as e:
        logging.error(f"[FTP] nlst failed: {e}")
        ftp.quit()
        return []

    downloaded = []
    for f in files:
        # only handle xml/xlsx (skip hidden & other files)
        if not (f.lower().endswith(".xml") or f.lower().endswith(".xlsx")):
            logging.info(f"[FTP] Skipping unsupported file: {f}")
            continue

        local_path = os.path.join(INPUT_FOLDER, f)
        os.makedirs(INPUT_FOLDER, exist_ok=True)

        logging.info(f"[FTP] Downloading: {f}")
        start = time.time()
        try:
            with open(local_path, "wb") as fp:
                ftp.retrbinary("RETR " + f, fp.write)
            elapsed = time.time() - start
            logging.info(f"[FTP] Completed: {f} in {int(elapsed*1000)} ms")
        except Exception as e:
            logging.error(f"[FTP] Download failed for {f}: {e}")
            continue

        # move to remote archive (safe-guard in try/except)
        try:
            ftp.rename(f"{REMOTE_INCOMING}/{f}", f"{REMOTE_ARCHIVE}/{f}")
            logging.info(f"[FTP] Remote archived: {f}")
        except Exception as e:
            logging.warning(f"[FTP] Could not archive remote {f}: {e}")

        downloaded.append(f)

    try:
        ftp.quit()
    except Exception:
        pass

    return downloaded


# =======================================================
# BATCH INSERT HELPER (returns number of rows inserted)
# =======================================================
def insert_batch_and_count(cursor, insert_query, batch_data):
    """
    Execute execute_values for batch_data and return how many rows were submitted.
    We count rows by len(batch_data) rather than depending on RETURNING/fetchall,
    which is more robust for logging counts.
    """
    if not batch_data:
        return 0

    # make a copy of length
    n = len(batch_data)
    execute_values(cursor, insert_query, batch_data)
    # clear the original batch
    batch_data.clear()
    return n


# =======================================================
# PROCESS XML
# =======================================================
def process_xml(file_path, cursor, batch_size=BATCH_SIZE):
    logging.info(f"[XML] Processing: {file_path}")
    tree = ET.parse(file_path)
    root = tree.getroot()

    batch_data = []
    total_inserted = 0

    insert_query = """
        INSERT INTO sales_orders (sku, sale_qty, store_code, consolidate, is_read)
        VALUES %s
    """

    for order_line in root.findall("OrderLine"):
        sku_el = order_line.find("ArticleNumber")
        store_el = order_line.find("P1SlaveId")
        if sku_el is None:
            continue
        sku = sku_el.text.strip() if sku_el.text else ""
        store_code = store_el.text.strip() if (store_el is not None and store_el.text) else ""
        sale_qty = 1

        if not sku:
            continue

        batch_data.append((sku, sale_qty, store_code, 0, 0))
        if len(batch_data) >= batch_size:
            inserted = insert_batch_and_count(cursor, insert_query, batch_data)
            total_inserted += inserted
            logging.info(f"[XML] Batch inserted: {inserted} rows")

    # leftover
    if batch_data:
        inserted = insert_batch_and_count(cursor, insert_query, batch_data)
        total_inserted += inserted
        logging.info(f"[XML] Final batch inserted: {inserted} rows")

    logging.info(f"[XML] Total inserted for {os.path.basename(file_path)}: {total_inserted}")
    return total_inserted


# =======================================================
# PROCESS XLSX
# =======================================================
def process_xlsx(file_path, cursor, batch_size=BATCH_SIZE):
    logging.info(f"[XLSX] Processing: {file_path}")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.active

    # Read header row safely (first row)
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {name: idx for idx, name in enumerate(header_cells) if name is not None}

    if "Material" not in headers or "Slave ID" not in headers:
        raise Exception("XLSX missing required headers: Material, Slave ID")

    material_col = headers["Material"]
    slave_col = headers["Slave ID"]

    batch_data = []
    total_inserted = 0

    insert_query = """
        INSERT INTO sales_orders (sku, sale_qty, store_code, consolidate, is_read)
        VALUES %s
    """

    # iterate rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # guard against short rows
        sku = None
        store_code = ""
        if material_col < len(row):
            sku = row[material_col]
        if slave_col < len(row):
            store_code = row[slave_col]

        if sku is None:
            continue

        sku_str = str(sku).strip()
        store_str = str(store_code).strip() if store_code is not None else ""

        if not sku_str:
            continue

        batch_data.append((sku_str, 1, store_str, 0, 0))

        if len(batch_data) >= batch_size:
            inserted = insert_batch_and_count(cursor, insert_query, batch_data)
            total_inserted += inserted
            logging.info(f"[XLSX] Batch inserted: {inserted} rows")

    # leftover
    if batch_data:
        inserted = insert_batch_and_count(cursor, insert_query, batch_data)
        total_inserted += inserted
        logging.info(f"[XLSX] Final batch inserted: {inserted} rows")

    wb.close()
    logging.info(f"[XLSX] Total inserted for {os.path.basename(file_path)}: {total_inserted}")
    return total_inserted


# =======================================================
# MAIN ORDER PROCESSOR
# =======================================================
def orderProcess():

    logging.info("Starting FTP download (if available)")
    # -------- FTP DOWNLOAD FIRST (non-fatal)
    try:
        ftp_download_files()
    except Exception as e:
        logging.warning(f"FTP download step failed or skipped: {e}")

    conn = get_connection()
    if not conn:
        print(" DB connection failed<br>")
        logging.error("DB connection failed")
        return False, 0

    cursor = conn.cursor()

    os.makedirs(ARCHIVE_FOLDER, exist_ok=True)

    valid_files = [
        f for f in os.listdir(INPUT_FOLDER)
        if f.lower().endswith(".xml") or f.lower().endswith(".xlsx")
    ]

    if not valid_files:
        print(" No XML or XLSX files found.<br>")
        logging.info("No files found to process.")
        cursor.close()
        conn.close()
        return False, 0

    for filename in valid_files:
        file_path = os.path.join(INPUT_FOLDER, filename)
        logging.info(f"[PROCESS] Starting: {filename}")
        print(f" Processing File: {filename}")

        try:
            total_inserted = 0

            if filename.lower().endswith(".xml"):
                total_inserted = process_xml(file_path, cursor)

            elif filename.lower().endswith(".xlsx"):
                total_inserted = process_xlsx(file_path, cursor)

            # commit once per file
            conn.commit()
            logging.info(f"[COMMIT] {filename} committed — total rows inserted: {total_inserted}")
            print(f" Inserted rows: {total_inserted} <br>")

            shutil.move(file_path, os.path.join(ARCHIVE_FOLDER, filename))
            logging.info(f"[ARCHIVE] {filename} moved to archive")
            print(f" Moved to archive: {filename} <br>")

        except Exception as e:
            conn.rollback()
            logging.error(f"[ERROR] {filename}: {str(e)}")
            print(f" Error processing {filename}: {str(e)} <br>")

    # -------- RUN PROCEDURE AFTER ALL FILES --------
    try:
        logging.info("Running procedure: process_sales_orders()")
        cursor.execute("CALL process_sales_orders()")
        conn.commit()
        logging.info("Procedure process_sales_orders() executed successfully")
    except Exception as e:
        conn.rollback()
        logging.error(f"Procedure Error: {str(e)}")
        print(f" Error running procedure: {str(e)}")

    cursor.close()
    conn.close()

    logging.info("--------------- SALES ORDER IMPORT COMPLETED ---------------")
    print("All files processed! <br>")


orderProcess()
