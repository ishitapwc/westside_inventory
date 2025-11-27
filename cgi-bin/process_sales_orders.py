#!/usr/bin/env python3
print("Content-Type: text/html\n")

import os
import shutil
import xml.etree.ElementTree as ET
import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook
from db_connection import get_connection

INPUT_FOLDER = "/var/www/html/python/westside_inventory/orders"
ARCHIVE_FOLDER = "/var/www/html/python/westside_inventory/orders_archive"
BATCH_SIZE = 5000


def insert_batch(cursor, insert_query, batch_data, inserted_ids):
    """Insert using execute_values and collect new row IDs."""
    if batch_data:
        execute_values(cursor, insert_query, batch_data)
        rows = cursor.fetchall()
        inserted_ids.extend([row[0] for row in rows])
        batch_data.clear()


def process_xml(file_path, cursor):
    """Process XML and insert rows batch-wise."""
    tree = ET.parse(file_path)
    root = tree.getroot()

    batch_data = []
    inserted_ids = []

    insert_query = """
        INSERT INTO sales_orders (sku, sale_qty, store_code, consolidate, is_read)
        VALUES %s
        RETURNING id
    """

    for order_line in root.findall("OrderLine"):
        sku = order_line.find("ArticleNumber").text
        store_code = order_line.find("P1SlaveId").text
        sale_qty = 1

        batch_data.append((sku, sale_qty, store_code, 0, 0))

        if len(batch_data) >= BATCH_SIZE:
            insert_batch(cursor, insert_query, batch_data, inserted_ids)

    insert_batch(cursor, insert_query, batch_data, inserted_ids)

    return inserted_ids


def process_xlsx(file_path, cursor):
    """Process XLSX and insert rows batch-wise using headers Material + Slave ID."""
    wb = load_workbook(file_path)
    sheet = wb.active

    # Read header row
    headers = {cell.value: idx for idx, cell in enumerate(sheet[1])}

    # Ensure required columns exist
    if "Material" not in headers or "Slave ID" not in headers:
        raise Exception("XLSX missing required headers: Material, Slave ID")

    material_col = headers["Material"]
    slave_col = headers["Slave ID"]

    batch_data = []
    inserted_ids = []

    insert_query = """
        INSERT INTO sales_orders (sku, sale_qty, store_code, consolidate, is_read)
        VALUES %s
        RETURNING id
    """

    # Loop rows starting from row 2
    for row in sheet.iter_rows(min_row=2, values_only=True):
        sku = row[material_col]          # Material → sku
        store_code = row[slave_col]      # Slave ID → store_code
        sale_qty = 1                     # always 1

        if sku:
            batch_data.append((str(sku), sale_qty, str(store_code), 0, 0))

        if len(batch_data) >= BATCH_SIZE:
            insert_batch(cursor, insert_query, batch_data, inserted_ids)

    # Insert leftover rows
    insert_batch(cursor, insert_query, batch_data, inserted_ids)

    return inserted_ids


def orderProcess():
    conn = get_connection()
    if not conn:
        print(" DB connection failed<br>")
        return False, 0
    cursor = conn.cursor()

    os.makedirs(ARCHIVE_FOLDER, exist_ok=True)

    # ---------- CHECK FOR EMPTY FOLDER ----------
    valid_files = [
        f for f in os.listdir(INPUT_FOLDER)
        if f.endswith(".xml") or f.endswith(".xlsx")
    ]

    if not valid_files:
        print(" No XML or XLSX files found in the input folder.<br>")
        cursor.close()
        conn.close()
        return False, 0
    # --------------------------------------------

    for filename in os.listdir(INPUT_FOLDER):
        file_path = os.path.join(INPUT_FOLDER, filename)

        if not (filename.endswith(".xml") or filename.endswith(".xlsx")):
            continue

        print(f" Processing File: {filename}")

        try:
            inserted_ids = []

            # XML File
            if filename.endswith(".xml"):
                inserted_ids = process_xml(file_path, cursor)

            # XLSX File
            elif filename.endswith(".xlsx"):
                inserted_ids = process_xlsx(file_path, cursor)

            # Call stored procedure
            # cursor.execute("CALL sales_orders_total()")

            # Mark rows as consolidated
            # if inserted_ids:
            #     cursor.execute(
            #         "UPDATE sales_orders SET consolidate = 1 WHERE id = ANY(%s)",
            #         (inserted_ids,)
            #     )

            conn.commit()

            shutil.move(file_path, os.path.join(ARCHIVE_FOLDER, filename))
            print(f" Moved to archive: {filename} <br>")

        except Exception as e:
            print(f" Error processing {filename}: {str(e)} <br>")

    # --------- RUN PROCEDURE ONLY ONCE AFTER ALL FILES ---------
    try:
        print(" Running process_sales_orders()...")
        cursor.execute("CALL process_sales_orders()")
        conn.commit()
    except Exception as e:
        print(f" Error running procedure: {str(e)}")
    # -----------------------------------------------------------

    

    cursor.close()
    conn.close()

    print("All files processed! <br>")


orderProcess()